# tools/excel_tools.py
"""
Excel and spreadsheet tools for the AI Tools framework
Includes reading, writing, and manipulating Excel files
"""

import os
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
import json
from core.base import BaseTool, ToolDefinition, ToolParameter, ToolResult, ToolResultType
from core.registry import registry

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

class ReadExcelTool(BaseTool):
    """Read data from Excel files"""
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
    
    @property
    def definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="read_excel",
            description="Read data from Excel files (.xlsx, .xls) and return as JSON or CSV format",
            category="data_processing",
            parameters=[
                ToolParameter(
                    name="file_path",
                    description="Path to the Excel file to read",
                    param_type="string",
                    required=True
                ),
                ToolParameter(
                    name="sheet_name",
                    description="Name or index of the sheet to read (default: first sheet)",
                    param_type="string",
                    required=False
                ),
                ToolParameter(
                    name="output_format",
                    description="Output format for the data",
                    param_type="string",
                    required=False,
                    default="json"
                ),
                ToolParameter(
                    name="max_rows",
                    description="Maximum number of rows to read",
                    param_type="number",
                    required=False,
                    default=100
                ),
                ToolParameter(
                    name="has_header",
                    description="Whether the first row contains column headers",
                    param_type="boolean",
                    required=False,
                    default=True
                )
            ]
        )
    
    async def execute(self, file_path: str, sheet_name: Optional[str] = None,
                     output_format: str = "json", max_rows: int = 100,
                     has_header: bool = True) -> ToolResult:
        """Execute Excel reading"""
        if not PANDAS_AVAILABLE:
            return ToolResult(
                success=False,
                result_type=ToolResultType.ERROR,
                content="pandas library not available. Install with: pip install pandas openpyxl",
                error_message="pandas library not available"
            )
        
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                return ToolResult(
                    success=False,
                    result_type=ToolResultType.ERROR,
                    content=f"File not found: {file_path}",
                    error_message="File not found"
                )
            
            # Read Excel file
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=max_rows,
                header=0 if has_header else None
            )
            
            # Convert to requested format
            if output_format.lower() == "json":
                data = df.to_json(orient='records', indent=2)
                result_data = json.loads(data)
                content = json.dumps(result_data, indent=2)
            elif output_format.lower() == "csv":
                content = df.to_csv(index=False)
            else:
                content = str(df)
            
            # Get sheet info
            sheet_info = ""
            if EXCEL_AVAILABLE:
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    sheet_names = wb.sheetnames
                    sheet_info = f"Available sheets: {', '.join(sheet_names)}\n"
                    wb.close()
                except:
                    pass
            
            result_content = f"Successfully read Excel file: {file_path}\n"
            result_content += sheet_info
            result_content += f"Rows read: {len(df)}\n"
            result_content += f"Columns: {len(df.columns)}\n"
            if has_header:
                result_content += f"Column names: {', '.join(df.columns)}\n"
            result_content += f"\nData ({output_format} format):\n{content}"
            
            return ToolResult(
                success=True,
                content=result_content,
                result_type=ToolResultType.TEXT,
                metadata={
                    "tool": "read_excel",
                    "file_path": file_path,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "output_format": output_format
                }
            )
            
        except Exception as e:
            return ToolResult(
                success=False,
                result_type=ToolResultType.ERROR,
                content=f"Error reading Excel file: {str(e)}",
                error_message=f"Error reading Excel file: {str(e)}"
            )

class WriteExcelTool(BaseTool):
    """Create and write Excel files with formatting"""
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
    
    @property
    def definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="write_excel",
            description="Create Excel files with data, formatting, and charts - creates new or overwrites existing",
            category="data_processing",
            parameters=[
                ToolParameter(
                    name="file_path",
                    description="Path to the Excel file to create",
                    param_type="string",
                    required=True
                ),
                ToolParameter(
                    name="data",
                    description="Data as JSON string (array of objects or array of arrays)",
                    param_type="string",
                    required=True
                ),
                ToolParameter(
                    name="sheet_name",
                    description="Name of the worksheet",
                    param_type="string",
                    required=False,
                    default="Sheet1"
                ),
                ToolParameter(
                    name="headers",
                    description="Column headers as comma-separated string (if data is array of arrays)",
                    param_type="string",
                    required=False
                ),
                ToolParameter(
                    name="format_header",
                    description="Whether to format the header row with bold text and background",
                    param_type="boolean",
                    required=False,
                    default=True
                ),
                ToolParameter(
                    name="auto_width",
                    description="Whether to auto-adjust column widths",
                    param_type="boolean",
                    required=False,
                    default=True
                ),
                ToolParameter(
                    name="freeze_header",
                    description="Whether to freeze the header row",
                    param_type="boolean",
                    required=False,
                    default=True
                )
            ]
        )
    
    async def execute(self, file_path: str, data: str, sheet_name: str = "Sheet1",
                     headers: Optional[str] = None, format_header: bool = True,
                     auto_width: bool = True, freeze_header: bool = True) -> ToolResult:
        """Execute Excel creation"""
        if not EXCEL_AVAILABLE:
            return ToolResult(
                success=False,
                result_type=ToolResultType.ERROR,
                content="openpyxl library not available. Install with: pip install openpyxl",
                error_message="openpyxl library not available"
            )
        
        try:
            # Parse data
            try:
                parsed_data = json.loads(data)
            except json.JSONDecodeError:
                return ToolResult(
                    success=False,
                    result_type=ToolResultType.ERROR,
                    content="Invalid JSON format in data parameter",
                    error_message="Invalid JSON format in data"
                )
            
            # Create directory if needed
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Handle different data formats
            if isinstance(parsed_data, list) and len(parsed_data) > 0:
                if isinstance(parsed_data[0], dict):
                    # Array of objects - use keys as headers
                    header_row = list(parsed_data[0].keys())
                    ws.append(header_row)
                    
                    for row_data in parsed_data:
                        row_values = [row_data.get(key, "") for key in header_row]
                        ws.append(row_values)
                
                elif isinstance(parsed_data[0], list):
                    # Array of arrays
                    if headers:
                        header_row = [h.strip() for h in headers.split(',')]
                        ws.append(header_row)
                    
                    for row_data in parsed_data:
                        ws.append(row_data)
                
                else:
                    # Simple array - create single column
                    if headers:
                        ws.append([headers])
                    for item in parsed_data:
                        ws.append([item])
            
            # Apply formatting
            if format_header and ws.max_row > 0:
                # Format header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Auto-adjust column widths
            if auto_width:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            if freeze_header and ws.max_row > 1:
                ws.freeze_panes = "A2"
            
            # Add borders to data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Save the file
            wb.save(file_path)
            
            # Get file info
            file_size = os.path.getsize(file_path)
            
            result_content = f"Successfully created Excel file: {file_path}\n"
            result_content += f"Sheet name: {sheet_name}\n"
            result_content += f"Rows: {ws.max_row}\n"
            result_content += f"Columns: {ws.max_column}\n"
            result_content += f"File size: {file_size:,} bytes\n"
            result_content += f"Features applied: "
            
            features = []
            if format_header:
                features.append("header formatting")
            if auto_width:
                features.append("auto column width")
            if freeze_header:
                features.append("frozen header")
            
            result_content += ", ".join(features) if features else "none"
            
            return ToolResult(
                success=True,
                content=result_content,
                result_type=ToolResultType.TEXT,
                metadata={
                    "tool": "write_excel",
                    "file_path": file_path,
                    "rows": ws.max_row,
                    "columns": ws.max_column,
                    "file_size": file_size
                }
            )
            
        except Exception as e:
            return ToolResult(
                success=False,
                result_type=ToolResultType.ERROR,
                content=f"Error creating Excel file: {str(e)}",
                error_message=f"Error creating Excel file: {str(e)}"
            )

# Only register tools if libraries are available
if PANDAS_AVAILABLE:
    registry.register(ReadExcelTool)

if EXCEL_AVAILABLE:
    registry.register(WriteExcelTool)
